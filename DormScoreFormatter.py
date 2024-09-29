import logging
import os
import sys
import pandas as pd
import argparse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import re

def load_and_process_csv_files(folder_path):
    all_data = []
    for file in os.listdir(folder_path):
        if file.startswith('WeekScoreManage_') and file.endswith('.csv'):
            file_path = os.path.join(folder_path, file)
            df = pd.read_csv(file_path, encoding='gbk')
            all_data.append(df)
    
    combined_df = pd.concat(all_data, ignore_index=True)
    processed_df = combined_df[['楼号', '周', '房间', '床位', '总分', '整改意见']].copy()
    processed_df = processed_df.drop_duplicates(subset=['楼号', '房间', '床位'], keep='first')
    processed_df['整改意见'] = processed_df['整改意见'].apply(lambda x: re.sub(r'[^\w\s]', '', str(x)))
    processed_df = processed_df.sort_values(['房间', '床位'])
    
    return processed_df

def create_excel_file(df, output_file, email_prefix, folder_path):
    no_error = True
    wb = Workbook()
    ws = wb.active

    # Set column widths
    for col in ['A', 'B', 'C', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 6
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['H'].width = 25
    # Add title and header
    title = f"{df['楼号'].iloc[0]}{df['周'].iloc[0]}"
    ws.merge_cells('A1:H1')
    ws['A1'] = title
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:H2')
    ws['A2'] = "勤工大队楼层长分队统一意见邮箱：thu.lczh@gmail.com"
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A3:H3')
    ws['A3'] = f'如有疑问请联系学生楼长：{email_prefix}@mails.tsinghua.edu.cn 或登陆家园网查询具体成绩'
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

    # Add data
    headers = ['房间', '床位', '总分', '整改意见']
    ROW_PER_PAGE = 55
    row = 5
    col = 1
    page = 1
    rows_on_page = 5

    for i, header in enumerate('ABCDEFGH'):
        ws[f'{header}4'] = headers[i % 4]
        ws[f'{header}4'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{header}4'].font = Font(name='宋体', size=11)

    for _, data_row in df.iterrows():
        if rows_on_page == ROW_PER_PAGE + 1:
            row = 5 if page == 1 else (page - 1) * ROW_PER_PAGE + 1
            col += 4
            rows_on_page = 5 if page == 1 else 1
            if col > 5:
                page += 1
                row = (page - 1) * ROW_PER_PAGE + 1
                col = 1
                rows_on_page = 1

        for i, header in enumerate(headers):
            if data_row[header] == '' or pd.isna(data_row[header]) or data_row[header] == 'nan':
                ws.cell(row=row, column=col+i, value='').fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
                logging.warning(f"Empty cell at row {row}, column {col+i}: {header}: {data_row[header]}")
                no_error = False
            else:
                ws.cell(row=row, column=col+i, value=data_row[header])
            ws.cell(row=row, column=col+i).alignment = Alignment(horizontal='center', vertical='center')
            font_size = 11 - max(0, (len(str(data_row[header])) - 10) // 2)
            ws.cell(row=row, column=col+i).font = Font(name='宋体', size=font_size)
        row += 1
        rows_on_page += 1

    # Set border for all cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    
    wb.save(os.path.join(folder_path, output_file))
    
    return no_error
    

def generatePDF(folder_path, excel_file):
    print("Generating PDF file...")
    import win32com.client

    excel = win32com.client.Dispatch('Excel.Application')
    excel.Visible = False
    try:
        wb = excel.Workbooks.Open(os.path.join(
            os.path.abspath(folder_path), excel_file))
        ws = wb.Worksheets[0]
        ws.PageSetup.HeaderMargin = excel.CentimetersToPoints(0.8)
        ws.PageSetup.FooterMargin = excel.CentimetersToPoints(0.8)
        ws.PageSetup.TopMargin = excel.CentimetersToPoints(1.6)
        ws.PageSetup.BottomMargin = excel.CentimetersToPoints(1.6)
        ws.PageSetup.FitToPagesWide = 1
        ws.PageSetup.CenterHorizontally = True
        pdf_file = os.path.splitext(excel_file)[0] + '.pdf'
        wb.ExportAsFixedFormat(0, os.path.join(os.path.abspath(folder_path), pdf_file))
        wb.Close(SaveChanges=True)
        print(f"PDF file '{pdf_file}' has been created successfully.")
    except Exception as e:
        logging.error('Error occurred while generating PDF file.', e)
    finally:
        excel.Quit()

def main():
    parser = argparse.ArgumentParser(description='Process WeekScoreManage CSV files and create Excel and PDF report.')
    parser.add_argument('--folder', default='.', help='path to the folder containing CSV files')
    parser.add_argument('--email', default='xxx', help='prefix of your THU email address, like \'yunyang-21\'')
    parser.add_argument('--pdf', default='true', help='whether to generate PDF file', choices=['true', 'false'])
    parser.add_argument('--overwrite', default='false', help='overwrite existing excel files', choices=['true', 'false'])
    parser.add_argument('--pdfOnly', default='false', help='generate PDF file only', choices=['true', 'false'])
    parser.add_argument('--clean', default='false', help='clean up the csv files after success', choices=['true', 'false'])
    args = parser.parse_args()
    
    no_error = True
    
    if args.pdfOnly.lower() == 'true':
        for file in os.listdir(args.folder):
            if file.endswith('.xlsx'):
                generatePDF(args.folder, file)
        return

    df = load_and_process_csv_files(args.folder)
    
    # Check if the output file already exists
    output_file = f"{df['楼号'].iloc[0]}{df['周'].iloc[0]}.xlsx"
    if os.path.exists(os.path.join(args.folder, output_file)) and args.overwrite.lower() == 'false':
        print(f"Excel file '{output_file}' already exists. Please set '--overwrite true' to overwrite it.")
    else:
        no_error = create_excel_file(df, output_file, args.email, args.folder)
    print(f"Excel file '{output_file}' has been created successfully.")
    
    if args.pdf.lower() == 'true' and no_error:
        generatePDF(args.folder, output_file)
    elif not no_error:
        confirm = input("There are empty cells in the Excel file. Do you still want to generate PDF file? (Y/[N]) ")
        if confirm.lower() == 'y':
            generatePDF(args.folder, output_file)
        else:
            print("PDF file will not be generated.")
            
    if args.clean.lower() == 'true' and no_error:
        clean_up(args.folder)
    elif not no_error:
        confirm = input("There are empty cells in the Excel file. Do you still want to clean up the CSV files? (Y/[N]) ")
        if confirm.lower() == 'y':
            clean_up(args.folder)
            print("CSV files have been cleaned up.")
        else:
            print("CSV files will not be cleaned up.")
            
def clean_up(folder_path):
    print("Cleaning up CSV files...")
    for file in os.listdir(folder_path):
        if file.startswith('WeekScoreManage_') and file.endswith('.csv'):
            os.remove(os.path.join(folder_path, file))
    print("CSV files have been cleaned up.")

if __name__ == '__main__':
    sys.path.append(os.path.dirname(os.path.abspath(__file__)))
    main()
    os.system("pause")
